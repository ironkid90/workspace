from typing import List, Any, Dict, Optional

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import os
import time

from ..config import EXCEL_LOCK_TIMEOUT_S
from .common.pathing import resolve_in_allowed_base
from .common.results import error, from_exception, success

LOCK_SUFFIX = ".mcp.lock"


def _acquire_lock(path: str, timeout: int = EXCEL_LOCK_TIMEOUT_S) -> str:
    lock_path = f"{path}{LOCK_SUFFIX}"
    start = time.time()
    while True:
        try:
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_RDWR)
            os.close(fd)
            return lock_path
        except FileExistsError:
            if time.time() - start > timeout:
                raise TimeoutError("Could not acquire lock on workbook")
            time.sleep(0.1)


def inspect(workbook_path: str) -> Dict[str, Any]:
    lock_path = None
    try:
        p = resolve_in_allowed_base(workbook_path)
        if not p.exists():
            return error("not_found", f"Workbook not found: {workbook_path}")
        lock_path = _acquire_lock(str(p))
        wb = load_workbook(p, read_only=True, data_only=True)
        sheets = [{"name": ws.title, "max_row": ws.max_row, "max_column": ws.max_column} for ws in wb.worksheets]
        named_ranges = [nr.name for nr in wb.defined_names.definedName if nr.name]
        tables = []
        for ws in wb.worksheets:
            for tbl in getattr(ws, "tables", {}).values():
                tables.append({"sheet": ws.title, "name": tbl.name, "ref": tbl.ref})
        return success(sheets=sheets, named_ranges=named_ranges, tables=tables)
    except Exception as exc:
        return from_exception(exc)
    finally:
        if lock_path and os.path.exists(lock_path):
            os.remove(lock_path)


def read_range(workbook_path: str, sheet: str, a1_range: str, top_n: Optional[int] = None) -> Dict[str, Any]:
    lock_path = None
    try:
        p = resolve_in_allowed_base(workbook_path)
        if not p.exists():
            return error("not_found", f"Workbook not found: {workbook_path}")
        lock_path = _acquire_lock(str(p))
        wb = load_workbook(p, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            return error("not_found", f"Sheet not found: {sheet}")
        ws = wb[sheet]
        min_col, min_row, max_col, max_row = range_boundaries(a1_range)
        rows: List[List[Any]] = []
        for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            rows.append([None if v is None else v for v in r])
            if top_n and len(rows) >= top_n:
                break
        return success(rows=rows, range=a1_range)
    except Exception as exc:
        return from_exception(exc, default_code="invalid_path")
    finally:
        if lock_path and os.path.exists(lock_path):
            os.remove(lock_path)


def preview_write(workbook_path: str, sheet: str, a1_range: str, values: List[List[Any]]) -> Dict[str, Any]:
    lock_path = None
    try:
        p = resolve_in_allowed_base(workbook_path)
        if not p.exists():
            return error("not_found", f"Workbook not found: {workbook_path}")
        lock_path = _acquire_lock(str(p))
        if not values or not isinstance(values, list) or not values[0]:
            return error("invalid_path", "Values must be a non-empty 2D array.")
        wb = load_workbook(p, read_only=False, data_only=False)
        if sheet not in wb.sheetnames:
            return error("not_found", f"Sheet not found: {sheet}")
        ws = wb[sheet]
        min_col, min_row, _, _ = range_boundaries(a1_range)
        before: List[List[Any]] = []
        max_row = min_row + len(values) - 1
        max_cols = max(len(row) for row in values)
        max_col = min_col + max_cols - 1
        for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            before.append([None if v is None else v for v in r])
        for r_idx, row in enumerate(values):
            for c_idx, val in enumerate(row):
                ws.cell(row=min_row + r_idx, column=min_col + c_idx).value = val
        after: List[List[Any]] = []
        for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            after.append([None if v is None else v for v in r])
        return success(before=before, after=after, range=a1_range)
    except Exception as exc:
        return from_exception(exc, default_code="invalid_path")
    finally:
        if lock_path and os.path.exists(lock_path):
            os.remove(lock_path)


def commit_write(workbook_path: str, sheet: str, a1_range: str, values: List[List[Any]]) -> Dict[str, Any]:
    lock_path = None
    try:
        p = resolve_in_allowed_base(workbook_path)
        if not p.exists():
            return error("not_found", f"Workbook not found: {workbook_path}")
        lock_path = _acquire_lock(str(p))
        if not values or not isinstance(values, list) or not values[0]:
            return error("invalid_path", "Values must be a non-empty 2D array.")
        wb = load_workbook(p, read_only=False, data_only=False)
        if sheet not in wb.sheetnames:
            return error("not_found", f"Sheet not found: {sheet}")
        ws = wb[sheet]
        min_col, min_row, _, _ = range_boundaries(a1_range)
        for r_idx, row in enumerate(values):
            for c_idx, val in enumerate(row):
                ws.cell(row=min_row + r_idx, column=min_col + c_idx).value = val
        wb.save(p)
        return success(message="committed", path=str(p))
    except Exception as exc:
        return from_exception(exc, default_code="invalid_path")
    finally:
        if lock_path and os.path.exists(lock_path):
            os.remove(lock_path)


def find(
    workbook_path: str,
    query: Any,
    sheet: Optional[str] = None,
    a1_range: Optional[str] = None,
    match_case: bool = False,
    exact: bool = False,
    limit: int = 100,
) -> Dict[str, Any]:
    if query is None:
        return error("invalid_path", "Query cannot be null.")

    lock_path = None
    try:
        p = resolve_in_allowed_base(workbook_path)
        if not p.exists():
            return error("not_found", f"Workbook not found: {workbook_path}")
        lock_path = _acquire_lock(str(p))
        wb = load_workbook(p, read_only=True, data_only=True)
        sheets = [sheet] if sheet else wb.sheetnames
        matches: List[Dict[str, Any]] = []
        truncated = False
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                return error("not_found", f"Sheet not found: {sheet_name}")
            ws = wb[sheet_name]
            if a1_range:
                min_col, min_row, max_col, max_row = range_boundaries(a1_range)
            else:
                min_row, min_col = 1, 1
                max_row, max_col = ws.max_row, ws.max_column
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=False):
                for cell in row:
                    val = cell.value
                    if val is None:
                        continue
                    if exact:
                        is_match = val == query
                    else:
                        val_text = str(val)
                        query_text = str(query)
                        if not match_case:
                            val_text = val_text.lower()
                            query_text = query_text.lower()
                        is_match = query_text in val_text
                    if is_match:
                        matches.append({"sheet": sheet_name, "cell": cell.coordinate, "value": val})
                        if len(matches) >= limit:
                            truncated = True
                            break
                if truncated:
                    break
            if truncated:
                break
        return success(matches=matches, truncated=truncated)
    except Exception as exc:
        return from_exception(exc, default_code="invalid_path")
    finally:
        if lock_path and os.path.exists(lock_path):
            os.remove(lock_path)
